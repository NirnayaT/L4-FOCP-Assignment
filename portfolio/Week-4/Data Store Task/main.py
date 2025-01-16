import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import Workbook

class DataStoreApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Store System")
        self.root.geometry("400x400")
        self.FILE_NAME = "class_work2.xlsx"
        
        self.initialize_ui()
        self.initialize_excel()
        
    def initialize_ui(self):
        self.ui_frame = tk.Frame(self.root)
        self.ui_frame.grid(row=1, column=0, padx=10, pady=10)
        
        # Create entry fields
        tk.Label(self.ui_frame, text="Name").grid(row=0, column=0, sticky='e', padx=5)
        self.entry_name = tk.Entry(self.ui_frame)
        self.entry_name.grid(row=0, column=1, padx=5)
        
        tk.Label(self.ui_frame, text="Age").grid(row=1, column=0, sticky='e', padx=5)
        self.entry_age = tk.Entry(self.ui_frame)
        self.entry_age.grid(row=1, column=1, padx=5)
        
        # Save button
        tk.Button(self.ui_frame, text="Save", command=self.save_data).grid(row=2, column=1, pady=10)
        
        # Table
        self.table = ttk.Treeview(self.ui_frame, columns=("Name", "Age"), show="headings")
        self.table.heading("Name", text="Name")
        self.table.heading("Age", text="Age")
        self.table.grid(row=3, column=0, columnspan=2, pady=10)
        
    def initialize_excel(self):
        try:
            self.workbook = openpyxl.load_workbook(self.FILE_NAME)
            self.worksheet = self.workbook.active
            self.load_existing_data()
        except FileNotFoundError:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.append(["Name", "Age"])
            self.workbook.save(self.FILE_NAME)
            
    def load_existing_data(self):
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            self.table.insert("", tk.END, values=row)
            
    def validate_input(self, name, age):
        if not name or not age:
            messagebox.showerror("Error", "Please fill in all fields")
            return False
        try:
            int(age)
            return True
        except ValueError:
            messagebox.showerror("Error", "Age must be a number")
            return False
            
    def save_data(self):
        name = self.entry_name.get().strip()
        age = self.entry_age.get().strip()
        
        if not self.validate_input(name, age):
            return
            
        try:
            self.table.insert("", tk.END, values=(name, age))
            self.worksheet.append((name, age))
            self.workbook.save(self.FILE_NAME)
            
            # Clear entries after successful save
            self.entry_name.delete(0, tk.END)
            self.entry_age.delete(0, tk.END)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")


root = tk.Tk()
app = DataStoreApp(root)
root.mainloop()
